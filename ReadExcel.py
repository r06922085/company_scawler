import csv
import requests
import os
import openpyxl

FILENAME = "CompanyData.csv"

def main():
	ReadCSV('CompanyData.csv')

def ReadCSV(FileName):
	writer = FileWriter('TaipeiData.xlsx')
	result = {}
	count = 0
	data_count = 0
	row_count = 0
	with open(FileName, newline='') as csvfile:
		rows = csv.reader(csvfile)
		for row in rows:
			row_count += 1
			if row_count >=245000 and row_count <245500:
				if isContinue(row):
					data_count += 1
					data = GetByNumber(row[0])
					if isDataContinue(data):
						count += 1
						writer.write(row, data)
				print(str(count) + "/" + str(data_count) + "/" + str(row_count))
	writer.close_writer()
		 
def isContinue(row):
	isContinue_ = True
	if "臺北市大安區" not in row[5] and "臺北市中山區" not in row[5] and "臺北市中正區" not in row[5] and "臺北市萬華區" not in row[5]:
		isContinue_ = False

	return isContinue_

def isDataContinue(data):
	isDataContinue_ = True
	try:
		capital = data['Capital_Stock_Amount']
		capital_real = data['Paid_In_Capital_Amount']
		boss_name = data['Responsible_Name']
		status = data['Company_Status_Desc']
		if capital_real != 0 and capital_real is not None:
			capital = capital_real
	except:
		capital = 0
		boss_name = ''
		status = ''
	if status != "核准設立":
		isDataContinue_ = False
	if capital < 6000000 or capital > 50000000:
		isDataContinue_ = False
	if boss_name == '':
		isDataContinue_ = False
	return isDataContinue_


def GetByNumber(businessAccountingNo):
	try:
	    result = requests.get("https://data.gcis.nat.gov.tw/od/data/api/5F64D864-61CB-4D0D-8AD9-492047CC1EA6?$format=json&$filter=Business_Accounting_NO eq " + businessAccountingNo + "&$skip=0&$top=50").json()[0]
	except:
		print("No such company number")
		result = None
	return result

class FileWriter():

	def __init__(self, recordFileName):
		self.recordFileName = recordFileName
		if os.path.isfile(recordFileName):
			#self.workbook = openpyxl.load_workbook(recordFileName)
			self.workbook = openpyxl.Workbook()
		else:
			self.workbook = openpyxl.Workbook()

		self.sheet = self.workbook.active
		self.sheet['A1'] = '統一編號'
		self.sheet['B1'] = '公司名稱'
		self.sheet['C1'] = '負責人'
		self.sheet['D1'] = '資本額'
		self.sheet['E1'] = '電話'
		self.sheet['F1'] = '成立日期'
		self.sheet['G1'] = '變更日期'
		self.sheet['H1'] = '地址'

	def get_result(self, row_data, website_data):
		result = {}
		result['Number'] = row_data[0]
		result['Company_Name'] = row_data[3]
		result['Boss_Name'] = website_data['Responsible_Name']
		result['capital'] = website_data['Paid_In_Capital_Amount']
		if result['capital'] == 0 or result['capital'] is None:
			result['capital'] = website_data['Capital_Stock_Amount']
		result['Phone'] = row_data[8]
		result['Address'] = row_data[5]
		result['Start'] = website_data['Company_Setup_Date']
		result['Change'] = website_data['Change_Of_Approval_Data']
		return result

	def write(self, row_data, website_data):
		result = self.get_result(row_data, website_data)
		result['capital'] = int(result['capital']/1000)
		result['capital'] = format(result['capital'], ',')
		self.sheet.append([  result['Number'],
						result['Company_Name'],
						result['Boss_Name'],
						result['capital'],
						result['Phone'],
						result['Address'],
						result['Start'],
						result['Change']])
	def close_writer(self):
		self.workbook.save(self.recordFileName)


main()
