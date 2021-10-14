import requests
import os
import openpyxl


def main():
	ReadEXCEL('動態電推.xlsx')

def ReadEXCEL(FileName):
	writer = FileWriter('動態電推_詳細.xlsx')
	result = {}
	count = 0
	data_count = 0
	row_count = 0
	rows = openpyxl.load_workbook(FileName)
	rows = rows['工作表1']
	for row in rows:
		number = row[0].value
		if number is None:
			continue
		data = GetByNumber(str(number))
		comment = ""
		writer.write(comment, data)
	writer.close_writer()

def GetByNumber(businessAccountingNo):
	try:
	    result = requests.get("https://data.gcis.nat.gov.tw/od/data/api/5F64D864-61CB-4D0D-8AD9-492047CC1EA6?$format=json&$filter=Business_Accounting_NO eq " + businessAccountingNo + "&$skip=0&$top=50").json()[0]
	except:
		try:
			businessAccountingNo = "0" + businessAccountingNo
			result = requests.get("https://data.gcis.nat.gov.tw/od/data/api/5F64D864-61CB-4D0D-8AD9-492047CC1EA6?$format=json&$filter=Business_Accounting_NO eq " + businessAccountingNo + "&$skip=0&$top=50").json()[0]
		except:
			print("No such company number")
			result = None
	return result

class FileWriter():

	def __init__(self, recordFileName):
		self.recordFileName = recordFileName
		if os.path.isfile(recordFileName):
			#self.workbook = openpyxl.load_workbook(recordFileName)
			self.workbook = openpyxl.Workbook()
		else:
			self.workbook = openpyxl.Workbook()

		self.sheet = self.workbook.active
		self.sheet['A1'] = '統一編號'
		self.sheet['B1'] = '公司名稱'
		self.sheet['C1'] = '負責人'
		self.sheet['D1'] = '資本額'
		self.sheet['E1'] = '電話'
		self.sheet['F1'] = '成立日期'
		self.sheet['G1'] = '變更日期'
		self.sheet['H1'] = '地址'

	def get_result(self, website_data):
		result = {}
		result['Number'] = website_data['Business_Accounting_NO']
		result['Company_Name'] = website_data['Company_Name']
		result['Boss_Name'] = website_data['Responsible_Name']
		if website_data['Paid_In_Capital_Amount'] != 0 and website_data['Paid_In_Capital_Amount'] is not None:
			result['capital'] = website_data['Paid_In_Capital_Amount']
		else:
			result['capital'] = website_data['Capital_Stock_Amount']
		result['Address'] = website_data['Company_Location']
		result['Start'] = website_data['Company_Setup_Date']
		result['Change'] = website_data['Change_Of_Approval_Data']
		return result

	def write(self, describe, website_data):
		try:
			result = self.get_result(website_data)
		except:
			return
		print(result['capital'])
		result['capital'] = int(result['capital']/1000)
		result['capital'] = format(result['capital'], ',')
		self.sheet.append([  result['Number'],
						result['Company_Name'],
						result['Boss_Name'],
						result['capital'],
						describe,
						result['Address'],
						result['Start'],
						result['Change']])
	def close_writer(self):
		self.workbook.save(self.recordFileName)


main()
