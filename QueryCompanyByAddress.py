# -*- coding: UTF-8 -*-

from bs4 import BeautifulSoup
import getopt
import openpyxl
import os.path
import requests
import sys
import time

request_headers = {
    'user-agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36',
    'referer':'http://findbiz.nat.gov.tw/fts/query/QueryList/queryList.do',
}

form_data = {
	'errorMsg': '',
	'validatorOpen': 'N',
	'rlPermit': '0',
	'userResp': '', 
	'curPage': '0',
	'fhl': 'zh_TW',
	'qryCond': '',
	'infoType': 'A',
	'qryType': 'cmpyType',
	'cmpyType': 'true',
	'brCmpyType': '',
	'busmType': '',
	'factType': '',
	'lmtdType': '',
	'isAlive': 'true',
	'busiItemMain': '', 
	'busiItemSub': '',
	'city': 'TPI_6300'
}

host = "https://findbiz.nat.gov.tw"

def main(argv):
	queryCondition, recordFileName, startPage = ParseParameter(argv)
	results = QueryCompanyDetail(queryCondition, recordFileName+'.xlsx', startPage)
	ExportResult(results, recordFileName+'.xlsx')

def ParseParameter(argv):
	try:
		opts, args = getopt.getopt(argv,"hc:o:p:",["condition=", "output=", "page="])
	except getopt.GetoptError:
		print("QueryCompanyDetail.py -c <QueryCondition> -o <OutputFileName> -p <StartPage>")
		sys.exit(2)
	for opt, arg in opts:
		if opt == '-h':
			print("QueryCompanyDetail.py -c <QueryCondition> -o <OutputFileName> -p <StartPage>")
			sys.exit()
		elif opt in ("-c", "--condition"):
			queryCondition = arg
		elif opt in ("-o", "--output"):
			outputFileName = arg
		elif opt in ("-p", "--page"):
			startPage = arg

	return queryCondition, outputFileName, startPage

def QueryCompanyDetail(queryCondition, recordFileName, startPage):
	form_data['qryCond'] = queryCondition

	results = list()
	businessAccountingNoSet = ReadExistingRecord(recordFileName)
	
	for i in range(10):
		print("It is %i attempt" % (i+1))
		totalPage = 1
		isSetTotalPage = False
		currentPage = int(startPage) - 1

		if currentPage >= totalPage:
			res = requests.post(
					host + "/fts/query/QueryList/queryList.do", 
					headers=request_headers, 
					data=form_data)
			res.encoding = 'utf8'
			soup = BeautifulSoup(res.text, "html.parser")

			if isSetTotalPage == False:
				if soup.find("input", id="totalPage") is not None:
					totalPage = int(soup.find("input", id="totalPage").get('value'))
				isSetTotalPage = True

		while currentPage < totalPage:
			form_data['curPage'] = str(currentPage)
			res = requests.post(
					host + "/fts/query/QueryList/queryList.do", 
					headers=request_headers, 
					data=form_data)
			res.encoding = 'utf8'
			soup = BeautifulSoup(res.text, "html.parser")

			if isSetTotalPage == False:
				if soup.find("input", id="totalPage") is not None:
					totalPage = int(soup.find("input", id="totalPage").get('value'))
				isSetTotalPage = True
			
			contentBlocks = soup.find_all("div", {"class", "panel panel-default"})

			if len(contentBlocks) == 0:
				print('Need your attention')
				return results

			for contentBlock in contentBlocks:
				try:
					content = contentBlock.find_all("div")[1].text
				except IndexError:
					print("No record.")
					break
				index = content.find("統一編號")
				businessAccountingNo = content[index+5:index+13]

				if businessAccountingNo not in businessAccountingNoSet:
					# Invoke API to obtain company detail by business account number
					try:
						results.append(requests.get("https://data.gcis.nat.gov.tw/od/data/api/5F64D864-61CB-4D0D-8AD9-492047CC1EA6?$format=json&$filter=Business_Accounting_NO eq " + businessAccountingNo + "&$skip=0&$top=50").json()[0])
						businessAccountingNoSet.add(businessAccountingNo)
					except ValueError:
						print("Current business Account Number is incorrect. Continue to parse next record.")
						continue

			currentPage += 1
			print(str(currentPage)+"/"+str(totalPage))
			time.sleep(2)
	
	return results

def ReadExistingRecord(recordFileName):

	existingRecord = set()
	if os.path.isfile(recordFileName):
		workbook = openpyxl.load_workbook(recordFileName)
		sheet = workbook.active

		for cell in sheet['A'][1:]:
			existingRecord.add(cell.value)

	return existingRecord

def ExportResult(results, recordFileName):

	if os.path.isfile(recordFileName):
		workbook = openpyxl.load_workbook(recordFileName)
	else:
		workbook = openpyxl.Workbook()

	sheet = workbook.active
	sheet['A1'] = '統一編號'
	sheet['B1'] = '公司名稱'
	sheet['C1'] = '資本總額'
	sheet['D1'] = '代表人姓名'
	sheet['E1'] = '公司所在地'
	sheet['F1'] = '核准設立日期'

	for result in results:
		if result['Capital_Stock_Amount'] < 1000000:
			continue
		result['Capital_Stock_Amount'] = int(result['Capital_Stock_Amount']/1000)
		result['Capital_Stock_Amount'] = format(result['Capital_Stock_Amount'], ',')
		sheet.append([  result['Business_Accounting_NO'],
						result['Company_Name'],
						result['Capital_Stock_Amount'],
						result['Responsible_Name'],
						result['Company_Location'],
						result['Company_Setup_Date']])

	workbook.save(recordFileName)

if __name__ == "__main__":
	main(sys.argv[1:])